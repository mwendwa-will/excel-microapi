from openpyxl import load_workbook
wb = load_workbook(filename = "hello_world.xlsx")

def print_rows():
         for row in sheet.iter_rows(values_only=True):
            print(row)


wb.sheetnames

sheet = wb.active


wb.save(filename = "hello_world.xlsx")
print_rows()